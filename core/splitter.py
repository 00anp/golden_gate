import os
import io
import datetime
import msoffcrypto
import openpyxl
from core.helpers import EXPORT_COLUMNS, col_letter_to_index, safe_str
from core.models import (
    DELIVERY_REQUIRES_PASSWORD,
    DELIVERY_SFTP,
    DELIVERY_SFTP_WITH_PASSWORD,
)

FOLDER_NAMES = {
    DELIVERY_REQUIRES_PASSWORD:  "Requires password",
    DELIVERY_SFTP:               "SFTP",
    DELIVERY_SFTP_WITH_PASSWORD: "SFTP with password",
}

DEFAULT_PASSWORD = "Glass2025!"


def get_unique_dsc_company(ws) -> list[str]:
    """Returns a sorted list of unique values from column A (DSC company)."""
    companies = set()
    for i in range(2, ws.max_row + 1):
        value = safe_str(ws.cell(i, 1).value)
        if value:
            companies.add(value)
    return sorted(companies)


def build_export_workbook(ws, company: str) -> openpyxl.Workbook:
    """Builds a new workbook containing only rows matching company,
    with only the EXPORT_COLUMNS columns."""
    col_indices = [col_letter_to_index(c) for c in EXPORT_COLUMNS]

    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # Write header row
    header_row = [ws.cell(1, idx).value for idx in col_indices]
    new_ws.append(header_row)

    # Write matching data rows
    for i in range(2, ws.max_row + 1):
        if safe_str(ws.cell(i, 1).value) == company:
            row_data = [ws.cell(i, idx).value for idx in col_indices]
            new_ws.append(row_data)

    return new_wb


def protect_xlsx_with_password(wb: openpyxl.Workbook, password: str) -> bytes:
    """Saves workbook to an in-memory buffer and encrypts it with msoffcrypto."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    encrypted = io.BytesIO()
    office_file = msoffcrypto.OfficeFile(buffer)
    office_file.encrypt(password, encrypted)
    return encrypted.getvalue()


def split_and_protect(
    processed_wb,
    output_folder: str,
    passwords: dict,
    progress_callback: callable,
    status_callback: callable,
) -> list[str]:
    """Splits the processed workbook by company, applies passwords,
    and routes each file to the correct delivery subfolder."""

    ws        = processed_wb.active
    timestamp = datetime.datetime.now().strftime("%Y_%m_%d_%H_%M")
    base_dir  = os.path.join(output_folder, timestamp)

    # Create all 3 subfolders upfront
    for folder_name in FOLDER_NAMES.values():
        os.makedirs(os.path.join(base_dir, folder_name), exist_ok=True)

    companies     = get_unique_dsc_company(ws)
    created_files = []
    progress      = 0.88

    for company in companies:
        step        = (1.0 - progress) / max(len(companies), 1)
        company_wb  = build_export_workbook(ws, company)
        filename    = f"{company}_NEW_GMC_{timestamp}.xlsx"
        company_cfg = passwords.get(company)

        # Determine delivery method (default: requires_password)
        if company_cfg is not None:
            method = company_cfg.delivery_method
        else:
            method = DELIVERY_REQUIRES_PASSWORD

        folder_name = FOLDER_NAMES[method]
        filepath    = os.path.join(base_dir, folder_name, filename)

        if method == DELIVERY_SFTP:
            # No password — save plain
            company_wb.save(filepath)

        elif method == DELIVERY_REQUIRES_PASSWORD:
            # Use configured password or DEFAULT_PASSWORD
            pwd       = (company_cfg.password if (company_cfg and company_cfg.password)
                         else DEFAULT_PASSWORD)
            protected = protect_xlsx_with_password(company_wb, pwd)
            with open(filepath, "wb") as f:
                f.write(protected)

        elif method == DELIVERY_SFTP_WITH_PASSWORD:
            # Password-protected but goes to SFTP folder
            pwd       = (company_cfg.password if (company_cfg and company_cfg.password)
                         else DEFAULT_PASSWORD)
            protected = protect_xlsx_with_password(company_wb, pwd)
            with open(filepath, "wb") as f:
                f.write(protected)

        created_files.append(filepath)
        progress += step
        progress_callback(progress)
        status_callback(f"{company} → {folder_name}")

    progress_callback(1.0)
    status_callback("Process complete.")
    return created_files