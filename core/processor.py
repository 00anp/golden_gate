import openpyxl
from core.rules_engine import load_rules, get_rule, get_payment_tier
from core.helpers import safe_float, safe_str, ceiling, col_letter_to_index, HEADERS


def apply_settlement_rules(ws, row:int, settlement_rules):
    gmc_customer_prefix:str = safe_str(ws.cell(row=row, column=17).value)
    balance:float = safe_float(ws.cell(row=row, column=12).value)
    current_pct:float = safe_float(ws.cell(row=row, column=26).value)

    rule = get_rule(settlement_rules, gmc_customer_prefix)
    handled:bool = False

    if rule is None:
        ws.cell(row=row, column=37).value = current_pct
        return
    
    if rule.review_column is not None:
        column = ws.cell(row=row, column=rule.review_column).value
        if column == rule.value_to_review:
            ws.cell(row=row, column=39).value = "X"
            return
        
    if rule.copy_z_to_ak:
        ws.cell(row=row, column=37).value = current_pct
        return
    
    if rule.z_greater_than_threshold:
        if current_pct > 0.45 and balance > rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_high
            ws.cell(row=row, column=37).value = rule.ak_high
        elif current_pct > 0.45 and balance <= rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_low
            ws.cell(row=row, column=37).value = rule.ak_low
        elif current_pct <= 0.45:
            pass
        handled = True

    if rule.z_lower_than_threshold:
        if current_pct < 0.45:
            ws.cell(row=row, column=37).value = current_pct
        elif current_pct == 0.45:
            if rule.balance_threshold > 0:
                ws.cell(row=row, column=37).value = rule.ak_high
            else:
                ws.cell(row=row, column=26).value = rule.z_high
                ws.cell(row=row, column=37).value = rule.ak_high
        elif current_pct > 0.45 and balance > rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_high
            ws.cell(row=row, column=37).value = rule.ak_high
        elif current_pct > 0.45 and balance <= rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_low
            ws.cell(row=row, column=37).value = rule.ak_low
        handled = True
    if not handled:
        if balance >= rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_high
            ws.cell(row=row, column=37).value = rule.ak_high
        elif balance < rule.balance_threshold:
            ws.cell(row=row, column=26).value = rule.z_low
            ws.cell(row=row, column=37).value = rule.ak_low
    
    if rule.mark_am:
        ws.cell(row=row, column=39).value = "X"
    
    if rule.mark_aq:
        ws.cell(row=row, column=43).value = "X"


def apply_payment_terms(ws, row:int, settlement, payment_tiers):
    gmc_customer_prefix:str = safe_str(ws.cell(row=row, column=17).value)
    tier = get_payment_tier(payment_tiers, settlement) 

    if tier is None:
        return

    if gmc_customer_prefix.startswith("LPL"):
        ws.cell(row=row, column=14).value = tier.min_payment
        ws.cell(row=row, column=15).value = tier.max_term_lpl
    else:
        ws.cell(row=row, column=14).value = tier.min_payment
        ws.cell(row=row, column=15).value = tier.max_term_default


def process_file(input_path:str, progress_callback=None, status_callback=None)-> tuple[openpyxl.Workbook, dict[str, int]]:

    def update(pct: float, msg: str):
        if status_callback:
            status_callback(msg)
        if progress_callback:
            progress_callback(pct)
    

    settlement_rules, payment_tiers = load_rules()
    update(0.0, "Opening file...")
    wb = openpyxl.load_workbook(input_path) 
    ws = wb.active

    rules_applied: dict[str, int] = {}

    update(0.05, "Renaming headers...")
    for col_letter, name in HEADERS.items():
        ws.cell(row=1, column=col_letter_to_index(col_letter)).value = name
    
    last_row = ws.max_row
    total_rows = max(last_row - 1, 1)

    update(0.10, f"Processing {total_rows} data rows...")

    for i in range(2, last_row + 1):
        col_a = safe_str(ws.cell(row=i, column=1).value)
        col_b = str(ws.cell(row=i, column=2).value or "")
        #Replace in col B according to col A
        if col_a in ("NDR", "LORG"):
            ws.cell(row=i, column=2).value = col_b.replace("DS", "DS-")
        elif col_a in ("BEYOND", "FLLG"):
            ws.cell(row=i, column=2).value = col_b.replace("P", "P-")
        #Copy col R to col AP
        ws.cell(row=i, column=42).value = ws.cell(row=i, column=18).value
        # Copy col AD to col AO
        ws.cell(row=i, column=41).value = ws.cell(row=i, column=30).value

        gmc_prefix = safe_str(ws.cell(row=i, column=17).value)
        rule = get_rule(settlement_rules, gmc_prefix)
        
        if rule is not None:
            rules_applied[rule.prefix] = rules_applied.get(rule.prefix, 0) + 1

        apply_settlement_rules(ws, i, settlement_rules)

        z_val  = safe_float(ws.cell(row=i, column=26).value)
        ak_val = safe_float(ws.cell(row=i, column=37).value)
        balance = safe_float(ws.cell(row=i, column=12).value)

        settlement = ceiling(balance * z_val)
        ws.cell(row=i, column=13).value = settlement

        lump_sum = ceiling(balance * ak_val)
        ws.cell(row=i, column=38).value = lump_sum

        apply_payment_terms(ws, i, settlement, payment_tiers)
        # Copy col AE to col AN
        ws.cell(row=i, column=40).value = ws.cell(row=i, column=31).value

        progress_pct = 0.10 + (0.70 * (i - 1) / total_rows)
        if i % 50 == 0 or i == last_row:
            update(progress_pct, f"Processing row {i} of {last_row}...")
    
    update(0.85, "Business logic applied successfully.")
    return wb, rules_applied
